# integrations/excel_writer.py
"""Generación del archivo Excel para el proveedor."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook

# Columnas en la plantilla: C=descripción, D=color, E=espesor, F=longitud, G=ancho, H=cantidad, I=veta
_FILA_INICIO = 14
_COL_DESCRIPCION = 3   # C
_COL_COLOR = 4         # D
_COL_ESPESOR = 5       # E
_COL_LONGITUD = 6     # F
_COL_ANCHO = 7        # G
_COL_CANTIDAD = 8     # H
_COL_VETA = 9         # I
_COL_NUMERO = 2       # B (enumera las piezas)


def write_excel(rows: list[dict], output_path: str) -> None:
    """
    Escribe las filas en un archivo Excel nuevo (sin plantilla).

    Args:
        rows: Lista de diccionarios; las claves son los encabezados.
        output_path: Ruta del archivo .xlsx a crear.

    Raises:
        ValueError: Si rows está vacío o output_path no es .xlsx.
    """
    if not rows:
        raise ValueError("No hay filas para escribir")

    path = Path(output_path)
    if path.suffix.lower() != ".xlsx":
        raise ValueError("La ruta de salida debe tener extensión .xlsx")

    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("No se pudo obtener la hoja activa")

    headers = list(rows[0].keys())
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, key in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(key))

    wb.save(output_path)


def write_excel_from_template(
    template_path: str,
    rows: list[dict],
    output_path: str,
) -> None:
    """
    Carga la plantilla, escribe los datos desde la fila indicada (C14 en adelante) y guarda en output_path.

    Espera en cada fila las claves: DESCRIPCION, COLOR, ESPESOR, LONGITUD, ANCHO, CANTIDAD, VETA.
    Escribe en columnas C a I; B se rellena con el número de pieza (1, 2, 3...).
    """
    if not rows:
        raise ValueError("No hay filas para escribir")

    t = Path(template_path)
    if not t.exists():
        raise FileNotFoundError(f"Plantilla no encontrada: {template_path}")

    out = Path(output_path)
    if out.suffix.lower() != ".xlsx":
        raise ValueError("La ruta de salida debe tener extensión .xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(template_path)
    ws = wb.active
    if ws is None:
        raise RuntimeError("No se pudo obtener la hoja activa de la plantilla")

    for i, row_data in enumerate(rows, start=0):
        fila = _FILA_INICIO + i
        ws.cell(row=fila, column=_COL_NUMERO, value=i + 1)
        ws.cell(row=fila, column=_COL_DESCRIPCION, value=row_data.get("DESCRIPCION", ""))
        ws.cell(row=fila, column=_COL_COLOR, value=row_data.get("COLOR", ""))
        ws.cell(row=fila, column=_COL_ESPESOR, value=row_data.get("ESPESOR", ""))
        ws.cell(row=fila, column=_COL_LONGITUD, value=row_data.get("LONGITUD", ""))
        ws.cell(row=fila, column=_COL_ANCHO, value=row_data.get("ANCHO", ""))
        ws.cell(row=fila, column=_COL_CANTIDAD, value=row_data.get("CANTIDAD", ""))
        ws.cell(row=fila, column=_COL_VETA, value=row_data.get("VETA", ""))

    wb.save(output_path)
